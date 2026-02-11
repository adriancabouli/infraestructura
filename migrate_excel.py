import re
import pandas as pd
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / 'excel' / 'datos.xlsx'

OUT_EXPEDIENTES_CSV = "expedientes.csv"
OUT_GESTIONES_CSV = "gestiones.csv"

def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        return re.sub(r"\s+", " ", v).strip()
    return v

def find_header_row(ws, max_scan=20):
    for r in range(1, max_scan+1):
        row=[norm(ws.cell(r,c).value) for c in range(1, 25)]
        joined="|".join([str(x) for x in row if x is not None])
        if any(x in (joined or "") for x in ["Exp.", "EXPTE", "EXPTE.", "Expte"]):
            if "AÑO" in (joined or "") or any(isinstance(x,str) and x.lower().startswith("año") for x in row):
                return r, row
    return None, None

def col_map(header_row):
    m={}
    for idx, val in enumerate(header_row, start=1):
        if not isinstance(val,str):
            continue
        v=val.lower()
        if v.startswith("exp") or "expte" in v:
            m['expte']=idx
        elif v.startswith("año") or v=="anio":
            m['anio']=idx
        elif "edificio" in v:
            m['edificio']=idx
        elif "carátula" in v or "caratula" in v or "referencia" in v:
            m['caratula']=idx
        elif "fecha de ingreso" in v:
            m['fecha_ingreso']=idx
        elif "última gestión" in v or "ultima gestion" in v:
            m['ultima_gestion']=idx
        elif "se giró a" in v or "se giro a" in v:
            m['se_giro_a']=idx
        elif "tipo" in v and "tramite" in v:
            m['tipo_tramite']=idx
        elif v=="fecha":
            m['fecha']=idx
        elif "etiqueta" in v:
            m['etiqueta']=idx
        elif "resol" in v:
            m['resolucion']=idx
        elif "dependencia" in v and "actual" in v:
            m['dependencia_actual']=idx
    return m

def is_valid_expv(expv):
    if expv is None: return False
    if isinstance(expv,(int,float)):
        return expv < 1e9 and expv != 0
    if isinstance(expv,str):
        s=expv.strip()
        if len(s) > 30:
            return False
        if s.upper() in ["S/N","SN","S. N."]:
            return True
        return bool(re.search(r"\d", s))
    return False

def expte_code_from(sheet_name, expv):
    digits_sheet=re.sub(r"\D","", sheet_name)
    if is_valid_expv(expv):
        if isinstance(expv,(int,float)):
            return str(int(expv))
        s=str(expv).strip()
        if s.upper() in ["S/N","SN","S. N."]:
            return digits_sheet or "SN"
        d=re.sub(r"\D","", s)
        return d or (digits_sheet or sheet_name.strip())
    return digits_sheet or sheet_name.strip()

def parse_history(ws, expte_code):
    # Busca "REGISTRO HISTORICO" y luego tabla FECHA / ULTIMA GESTION / SE GIRO A / DEP ACTUAL
    reg_row=None
    for r in range(1, 120):
        for c in range(1, 12):
            v=norm(ws.cell(r,c).value)
            if isinstance(v,str) and "REGISTRO" in v.upper():
                reg_row=r; break
        if reg_row: break
    if not reg_row:
        return []

    h2_row=None
    for r in range(reg_row, reg_row+12):
        row=[norm(ws.cell(r,c).value) for c in range(1, 25)]
        if any(isinstance(x,str) and x.upper()=="FECHA" for x in row):
            h2_row=r; h2=row; break
    if not h2_row:
        return []

    cmap2={}
    for idx,val in enumerate(h2, start=1):
        if not isinstance(val,str): continue
        v=val.lower()
        if v=="fecha": cmap2['fecha']=idx
        elif "última gestión" in v or "ultima gestion" in v: cmap2['gestion']=idx
        elif "se giró a" in v or "se giro a" in v: cmap2['se_giro_a']=idx
        elif "dependencia" in v and "actual" in v: cmap2['dependencia_actual']=idx

    hist=[]
    blank=0
    for r in range(h2_row+1, h2_row+400):
        f=norm(ws.cell(r,cmap2.get('fecha',1)).value)
        g=norm(ws.cell(r,cmap2.get('gestion',2)).value)
        if f is None and g is None:
            blank+=1
            if blank>=5: break
            continue
        blank=0
        hist.append({
            "expte_code": expte_code,
            "fecha": f,
            "gestion": g,
            "se_giro_a": norm(ws.cell(r,cmap2.get('se_giro_a',3)).value) if cmap2.get('se_giro_a') else None,
            "dependencia_actual": norm(ws.cell(r,cmap2.get('dependencia_actual',4)).value) if cmap2.get('dependencia_actual') else None,
        })
    return hist

def main():
    wb=openpyxl.load_workbook(XLSX_PATH, data_only=True)
    expedientes=[]
    gestiones=[]
    skipped=[]

    for sn in wb.sheetnames:
        if sn in ["INGRESO","PLANILLA","1","Hoja 3","Hoja 4"]:
            continue

        ws=wb[sn]
        hr, header=find_header_row(ws)
        main_row=None
        cmap=None

        if hr:
            cmap=col_map(header)
            # busca primera fila con año cargado
            for r in range(hr+1, hr+15):
                an=norm(ws.cell(r, cmap.get('anio',2)).value) if cmap.get('anio') else None
                if an is not None:
                    main_row=r
                    break

        expv=norm(ws.cell(main_row, cmap.get('expte',1)).value) if (main_row and cmap and cmap.get('expte')) else None
        expte=expte_code_from(sn, expv)

        if main_row and cmap:
            def get(field):
                c=cmap.get(field)
                return norm(ws.cell(main_row,c).value) if c else None
            expedientes.append({
                "expte_code": expte,
                "anio": get("anio"),
                "edificio": get("edificio"),
                "caratula": get("caratula"),
                "fecha_ingreso": get("fecha_ingreso"),
                "ultima_gestion": get("ultima_gestion"),
                "se_giro_a": get("se_giro_a"),
                "tipo_tramite": get("tipo_tramite"),
                "fecha": get("fecha"),
                "etiqueta": get("etiqueta"),
                "resolucion": get("resolucion"),
                "dependencia_actual": get("dependencia_actual"),
            })
        else:
            # hoja sin encabezado principal: solo historial
            expedientes.append({
                "expte_code": expte,
                "anio": None,
                "edificio": None,
                "caratula": None,
                "fecha_ingreso": None,
                "ultima_gestion": None,
                "se_giro_a": None,
                "tipo_tramite": None,
                "fecha": None,
                "etiqueta": None,
                "resolucion": None,
                "dependencia_actual": None,
            })

        gestiones.extend(parse_history(ws, expte))

    df_exp = pd.DataFrame(expedientes)
    df_gest = pd.DataFrame(gestiones)

    # --- Limpieza de fechas (convierte lo inválido a vacío) ---
    for col in ['fecha_ingreso', 'fecha']:
        if col in df_exp.columns:
            df_exp[col] = pd.to_datetime(df_exp[col], errors='coerce').dt.date

    if 'fecha' in df_gest.columns:
        df_gest['fecha'] = pd.to_datetime(df_gest['fecha'], errors='coerce').dt.date

    # --- Export ---
    df_exp.to_csv(OUT_EXPEDIENTES_CSV, index=False)
    df_gest.to_csv(OUT_GESTIONES_CSV, index=False)

    print("OK ->", OUT_EXPEDIENTES_CSV, OUT_GESTIONES_CSV)

if __name__ == "__main__":
    main()